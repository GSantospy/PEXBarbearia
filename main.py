from fastapi import FastAPI, Request, Form, Depends, HTTPException
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse
from fastapi import status
from datetime import datetime, time, timedelta, date
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os
import logging

app = FastAPI(docs_url="/documentacao-endpoint",
              redoc_url="/documentacao-sistema",
              openapi_url="/minha-openapi.json")
templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")

agendamentos = []
estoque = []
contas = []
usuarios = {}
loginUsuario = None
horarioInicial = datetime.strptime("09:00", "%H:%M").time()
horarioFinal = datetime.strptime("22:00", "%H:%M").time()
intervaloHoras = 1

logging.basicConfig(filename="logs.log",
                    format="%(asctime)s - %(levelname)s - %(message)s",
                    level=logging.INFO)


@app.middleware("http")
async def logRequests(request: Request, call_next):
    """
    Middleware para registrar detalhes das requisições HTTP e respostas geradas.
    Loga o método da requisição, a URL, o corpo da requisição (se houver) e o status da resposta.
    """
    logging.info(f"Recebida requisição: {request.method} {request.url}")
    if request.method in ["POST", "PUT", "DELETE", "PATCH", "GET"]:
        body = await request.body()
        logging.info(
            f"Corpo da requisição: {body.decode('utf-8') if body else 'Sem corpo'}"
        )

    response = await call_next(request)
    logging.info(f"Resposta gerada: status {response.status_code}")
    return response


def verificarLogin(request: Request):
    """
    Verifica se o usuário está logado, caso contrário, redireciona para a página de login.
    Lança uma exceção HTTPException se o login não for realizado.
    """
    global loginUsuario
    if loginUsuario is None:
        raise HTTPException(status_code=303,
                            detail="Redirecionando para login",
                            headers={"Location": "/login"})


def obterHorariosDisponiveis(data: str):
    """
    Obtém os horários disponíveis para agendamento em uma data específica.
    Verifica se o horário está ocupado e retorna os horários disponíveis.
    """
    dataSelecionada = datetime.strptime(data, "%Y-%m-%d").date()
    horaAtual = datetime.combine(dataSelecionada, horarioInicial)
    horariosDisponiveis = []
    hoje = datetime.now().date()
    horaAtualAgora = datetime.now().time()

    while horaAtual.time() <= horarioFinal:
        ocupado = any(agendamento["data"] == data
                      and agendamento["hora"] == horaAtual.strftime("%H:%M")
                      for agendamento in agendamentos)

        if not ocupado and (dataSelecionada != hoje
                            or horaAtual.time() > horaAtualAgora):
            horariosDisponiveis.append(horaAtual.strftime("%H:%M"))

        horaAtual += timedelta(hours=intervaloHoras)

    return horariosDisponiveis


@app.get("/", response_class=HTMLResponse, dependencies=[Depends(verificarLogin)])
async def index(request: Request):
    """
    Endpoint que exibe a página inicial com agendamentos do dia, estoque crítico e contas a vencer.
    """
    hoje = datetime.now().date()
    agendamentosDia = [
        ag for ag in agendamentos
        if datetime.strptime(ag["data"], "%Y-%m-%d").date() == hoje
    ]
    estoqueCritico = [prod for prod in estoque if int(prod["quantidade"]) < 5]
    contasVencer = [conta for conta in contas if conta["status"] in ["Ativa", "Atraso"]]

    agendamentosDia = [{
        "cliente": ag["cliente"],
        "servico": ag["servico"],
        "data": datetime.strptime(ag["data"], "%Y-%m-%d").strftime("%d/%m/%Y"),
        "hora": ag["hora"]
    } for ag in agendamentosDia]

    return templates.TemplateResponse(
        "index.html", {
            "request": request,
            "agendamentosDia": agendamentosDia,
            "estoqueCritico": estoqueCritico,
            "contasVencer": contasVencer
        })


@app.get("/login", response_class=HTMLResponse)
async def loginPagina(request: Request):
    """
    Endpoint que exibe a página de login.
    """
    return templates.TemplateResponse("login.html", {"request": request})


@app.post("/login")
async def realizarLogin(request: Request,
                        usuario: str = Form(...),
                        senha: str = Form(...)):
    """
    Endpoint que processa o login do usuário.
    """
    global loginUsuario
    if usuario in usuarios and usuarios[usuario]["senha"] == senha:
        loginUsuario = usuario
        return RedirectResponse("/", status_code=303)
    return templates.TemplateResponse("login.html", {
        "request": request,
        "error": "Usuário ou senha inválidos."
    })


@app.get("/cadastro", response_class=HTMLResponse)
async def cadastroPagina(request: Request):
    """
    Endpoint que exibe a página de cadastro.
    """
    return templates.TemplateResponse("cadastro.html", {"request": request})


@app.post("/cadastro")
async def realizarCadastro(request: Request,
                           nome: str = Form(...),
                           email: str = Form(...),
                           usuario: str = Form(...),
                           senha: str = Form(...)):
    """
    Endpoint que processa o cadastro de um novo usuário.
    """
    usuarios[usuario] = {
        "nome": nome,
        "email": email,
        "usuario": usuario,
        "senha": senha
    }
    return RedirectResponse("/login", status_code=303)


@app.post("/logout", dependencies=[Depends(verificarLogin)])
async def logout():
    """
    Endpoint para realizar o logout do usuário.
    """
    global loginUsuario
    loginUsuario = None
    return RedirectResponse("/login", status_code=303)


@app.get("/agendar", response_class=HTMLResponse, dependencies=[Depends(verificarLogin)])
async def agendar(request: Request, data: str = None):
    """
    Endpoint para exibir a página de agendamento, com horários disponíveis para a data selecionada.
    """
    data = data or datetime.now().strftime("%Y-%m-%d")
    horariosDisponiveis = obterHorariosDisponiveis(data)
    hoje = date.today().isoformat()
    return templates.TemplateResponse(
        "agendar.html", {
            "request": request,
            "data": data,
            "horariosDisponiveis": horariosDisponiveis,
            "hoje": hoje
        })


@app.post("/agendar", response_class=HTMLResponse, dependencies=[Depends(verificarLogin)])
async def realizarAgendamento(request: Request,
                              servico: str = Form(...),
                              data: str = Form(...),
                              hora: str = Form(...)):
    """
    Endpoint para processar o agendamento de um serviço para um cliente.
    """
    for agendamento in agendamentos:
        if agendamento["data"] == data and agendamento["hora"] == hora:
            return templates.TemplateResponse(
                "agendar.html", {
                    "request": request,
                    "data": data,
                    "horariosDisponiveis": obterHorariosDisponiveis(data),
                    "hoje": date.today().isoformat(),
                    "mensagem": "Horário indisponível para a data selecionada."
                })
    nomeCliente = usuarios[loginUsuario]["nome"]
    novoAgendamento = {
        "cliente": nomeCliente,
        "servico": servico,
        "data": data,
        "hora": hora,
        "situacao": "Ativo"
    }
    agendamentos.append(novoAgendamento)
    return RedirectResponse(url="/agendamentos", status_code=303)


@app.get("/agendamentos", response_class=HTMLResponse, dependencies=[Depends(verificarLogin)])
async def listarAgendamento(request: Request):
    """
    Exibe a lista de agendamentos existentes na página de agendamentos.

    A função formata os dados de agendamentos e os envia para a página de agendamentos, exibindo as informações
    de cliente, serviço, data e hora do agendamento.
    """
    agendamentosFormatados = [{
        "cliente": ag["cliente"],
        "servico": ag["servico"],
        "data": datetime.strptime(ag["data"], "%Y-%m-%d").strftime("%d/%m/%Y"),
        "hora": ag["hora"]
    } for ag in agendamentos]

    return templates.TemplateResponse("agendamentos.html", {
        "request": request,
        "agendamentos": agendamentosFormatados
    })


@app.get("/excluir_agendamento/{index}", response_class=HTMLResponse, dependencies=[Depends(verificarLogin)])
async def excluirAgendamento(index: int, request: Request):
    """
    Exclui um agendamento pelo índice na lista de agendamentos.

    A função tenta excluir o agendamento especificado pelo índice. Caso o índice seja inválido (fora do alcance),
    uma mensagem de erro é retornada.
    """
    try:
        agendamentos.pop(index)
        mensagem = "Agendamento excluído com sucesso."
    except IndexError:
        mensagem = "Agendamento não encontrado."

    return templates.TemplateResponse("agendamentos.html", {
        "request": request,
        "agendamentos": agendamentos,
        "error": mensagem
    })


@app.get("/alterar_agendamento", response_class=HTMLResponse, dependencies=[Depends(verificarLogin)])
async def alterarAgendamento(request: Request):
    """
    Exibe o formulário para alterar um agendamento.

    Essa rota exibe a página de alteração de agendamento onde o usuário pode modificar informações
    sobre um agendamento existente.
    """
    return templates.TemplateResponse("alterarAgendamento.html", {"request": request})


@app.get("/estoque", response_class=HTMLResponse, dependencies=[Depends(verificarLogin)])
async def listarEstoque(request: Request, mensagem: str = None):
    """
    Exibe a lista de produtos no estoque.

    A função formata os dados do estoque, incluindo nome, quantidade e validade dos produtos, e exibe
    essas informações na página de estoque. Uma mensagem opcional pode ser exibida, como confirmação de
    ações realizadas no estoque.
    """
    estoqueFormatado = [{
        "nome": item["nome"],
        "quantidade": item["quantidade"],
        "validade": datetime.strptime(item["validade"], "%Y-%m-%d").strftime("%d/%m/%Y")
        if item.get("validade") else None
    } for item in estoque]
    return templates.TemplateResponse("estoque.html", {
        "request": request,
        "estoque": estoqueFormatado,
        "mensagem": mensagem
    })


@app.post("/estoque", response_class=HTMLResponse, dependencies=[Depends(verificarLogin)])
async def cadastrarEstoque(request: Request,
                           nome: str = Form(...),
                           quantidade: str = Form(...),
                           validade: str = Form(...)):
    """
    Cadastra um novo produto no estoque.

    A função recebe os dados do novo produto (nome, quantidade, validade) através do formulário, valida
    a quantidade e adiciona o novo produto ao estoque.
    """
    try:
        quantidade_int = int(quantidade)
    except ValueError:
        quantidade_int = 0
    estoque.append({
        "nome": nome,
        "quantidade": quantidade_int,
        "validade": validade
    })
    return RedirectResponse("/estoque", status_code=303)


@app.get("/alterar_estoque", response_class=HTMLResponse, dependencies=[Depends(verificarLogin)])
async def alterarEstoque(request: Request):
    """
    Exibe o formulário para alterar um produto no estoque.

    Essa rota exibe a página de alteração de produto, onde o usuário pode modificar as informações
    de um produto existente no estoque.
    """
    return templates.TemplateResponse("alterarEstoque.html", {"request": request})


@app.get("/excluir_produto/{index}", response_class=HTMLResponse, dependencies=[Depends(verificarLogin)])
async def excluirProduto(index: int, request: Request):
    """
    Exclui um produto do estoque pelo índice na lista de estoque.

    A função tenta excluir o produto especificado pelo índice. Caso o índice seja inválido (fora do alcance),
    uma mensagem de erro é retornada.
    """
    try:
        estoque.pop(index)
        mensagem = "Produto excluído com sucesso."
    except IndexError:
        mensagem = "Produto não encontrado."

    return RedirectResponse(f"/estoque?mensagem={mensagem}", status_code=303)


@app.get("/contas", response_class=HTMLResponse, dependencies=[Depends(verificarLogin)])
async def contasPagina(request: Request):
    """
    Exibe a lista de contas a pagar na página de contas.

    A função formata as informações das contas, como descrição, valor, vencimento e status, e as exibe
    na página de contas.
    """
    contasFormatadas = [{
        "descricao": conta["descricao"],
        "valor": conta["valor"],
        "vencimento": datetime.strptime(conta["vencimento"], "%Y-%m-%d").strftime("%d/%m/%Y"),
        "status": conta["status"]
    } for conta in contas]

    return templates.TemplateResponse("contas.html", {
        "request": request,
        "contas": contasFormatadas
    })


@app.post("/contas", response_class=HTMLResponse, dependencies=[Depends(verificarLogin)])
async def cadastrarConta(request: Request,
                         descricao: str = Form(...),
                         valor: str = Form(...),
                         vencimento: str = Form(...)):
    """
    Cadastra uma nova conta a pagar.

    A função recebe os dados da nova conta (descrição, valor, vencimento) e adiciona a conta
    à lista de contas a pagar.
    """
    status = "Ativa"
    contas.append({
        "descricao": descricao,
        "valor": valor,
        "vencimento": vencimento,
        "status": status
    })
    return RedirectResponse("/contas", status_code=303)


@app.get("/alterar_conta", response_class=HTMLResponse, dependencies=[Depends(verificarLogin)])
async def alterarContas(request: Request):
    """
    Exibe o formulário para alterar uma conta.

    Essa rota exibe a página de alteração de conta, onde o usuário pode modificar informações
    sobre uma conta existente.
    """
    return templates.TemplateResponse("alterarContas.html", {"request": request})


@app.get("/excluir_conta/{index}", response_class=HTMLResponse, dependencies=[Depends(verificarLogin)])
async def excluirConta(index: int, request: Request):
    """
    Exclui uma conta da lista de contas a pagar.

    A função tenta excluir a conta especificada pelo índice. Caso o índice seja inválido (fora do alcance),
    uma mensagem de erro é retornada.
    """
    try:
        contas.pop(index)
        mensagem = "Conta excluída com sucesso."
    except IndexError:
        mensagem = "Conta não encontrada."

    return RedirectResponse(f"/contas?mensagem={mensagem}", status_code=303)


@app.get("/alterar_status_conta/{index}/{status}", response_class=HTMLResponse, dependencies=[Depends(verificarLogin)])
async def alterarStatusConta(index: int, status: str, request: Request):
    """
    Altera o status de uma conta a pagar.

    A função recebe o índice de uma conta e o novo status desejado (Atraso ou Paga). Caso o status seja
    válido, ele é atualizado na conta correspondente. Se o índice for inválido, ou se o status não for
    um dos permitidos, uma mensagem de erro será retornada.
    """
    try:
        if status in ["Atraso", "Paga"]:
            contas[index]["status"] = status
            mensagem = f"Status da conta alterado para {status}."
        else:
            mensagem = "Status inválido."
    except IndexError:
        mensagem = "Conta não encontrada."

    return RedirectResponse(f"/contas?mensagem={mensagem}", status_code=303)


@app.get("/perfil", response_class=HTMLResponse, dependencies=[Depends(verificarLogin)])
async def perfilPagina(request: Request):
    """
    Exibe a página de perfil do usuário com seus agendamentos.

    Filtra os agendamentos do usuário logado e os exibe na página de perfil.
    """
    agendamentosUsuarios = [
        agendamento for agendamento in agendamentos
        if agendamento["cliente"] == usuarios[loginUsuario]["usuario"]
    ]
    agendamentosUsuarios = [{
        "cliente": agendamento["cliente"],
        "servico": agendamento["servico"],
        "data": datetime.strptime(agendamento["data"], "%Y-%m-%d").strftime("%d/%m/%Y"),
        "hora": agendamento["hora"],
        "situacao": agendamento["situacao"]
    } for agendamento in agendamentosUsuarios]

    return templates.TemplateResponse(
        "perfil.html", {
            "request": request,
            "usuario": loginUsuario,
            "agendamentos": agendamentosUsuarios
        })


@app.get("/relatorio/excel", response_class=FileResponse, dependencies=[Depends(verificarLogin)])
async def gerarRelatorioExcel():
    """
    Gera um relatório em formato Excel contendo informações sobre agendamentos, estoque e contas.

    O relatório será gerado em três abas: Agendamentos, Estoque e Contas. Cada aba conterá
    as informações pertinentes, como cliente, serviço, data, hora, status, nome do produto,
    quantidade, validade, etc. O arquivo será salvo como "relatorio.xlsx" e enviado como resposta.
    """
    caminhoArquivo = "relatorio.xlsx"
    wb = Workbook()
    wbAgendamentos = wb.active
    wbAgendamentos.title = "Agendamentos"
    wbAgendamentos.append(["Cliente", "Serviço", "Data", "Hora", "Situação"])
    for agendamento in agendamentos:
        wbAgendamentos.append([
            agendamento["cliente"], agendamento["servico"],
            datetime.strptime(agendamento['data'],
                              '%Y-%m-%d').strftime('%d/%m/%Y'),
            agendamento["hora"], agendamento["situacao"]
        ])

    wbEstoque = wb.create_sheet(title="Estoque")
    wbEstoque.append(["Nome", "Quantidade", "Validade"])
    for produto in estoque:
        wbEstoque.append([
            produto["nome"],
            produto["quantidade"],
            datetime.strptime(produto['validade'], '%Y-%m-%d').strftime('%d/%m/%Y')
        ])

    wbContas = wb.create_sheet(title="Contas")
    wbContas.append(["Descrição", "Valor", "Vencimento", "Status"])
    for conta in contas:
        wbContas.append([
            conta["descricao"],
            conta["valor"],
            datetime.strptime(conta['vencimento'], '%Y-%m-%d').strftime('%d/%m/%Y'),
            conta["status"]
        ])

    wb.save(caminhoArquivo)

    return FileResponse(
        caminhoArquivo,
        media_type= "applicatin/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="Relatorio.xlsx")


@app.get("/relatorio/pdf", response_class=FileResponse, dependencies=[Depends(verificarLogin)])
async def gerarRelatorioPDF():
    """
    Gera um relatório em PDF com informações de agendamentos, estoque e contas.

    A função cria um arquivo PDF contendo as informações detalhadas dos agendamentos, produtos no estoque
    e contas a pagar. O relatório é formatado com informações como nome do cliente, serviço, data e hora
    de agendamento, quantidade de produtos, validade, descrição e status das contas, sendo gerado em formato PDF
    para download.
    """
    caminhoArquivo = "relatorio.pdf"
    c = canvas.Canvas(caminhoArquivo, pagesize=letter)
    largura, altura = letter
    margemSuperior = altura - 50
    margemInferior = 50
    espacamento = 20

    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, margemSuperior, "Relatório de Agendamentos, Estoque e Contas")
    c.setFont("Helvetica", 12)

    y = margemSuperior - 50

    def verificarEspaco(c, y, altura):
        if y < margemInferior:
            c.showPage()
            c.setFont("Helvetica", 12)
            return margemSuperior
        return y

    c.drawString(50, y, "Agendamentos")
    y -= espacamento
    for agendamento in agendamentos:
        c.drawString(
            50, y,
            f"{agendamento['cliente']} - {agendamento['servico']} - {datetime.strptime(agendamento['data'], '%Y-%m-%d').strftime('%d/%m/%Y')} {agendamento['hora']} - {agendamento['situacao']}"
        )
        y -= espacamento
        y = verificarEspaco(c, y, altura)

    y -= espacamento

    c.drawString(50, y, "Estoque")
    y -= espacamento
    for produto in estoque:
        c.drawString(
            50, y,
            f"{produto['nome']} - {produto['quantidade']} - Validade: {datetime.strptime(produto['validade'], '%Y-%m-%d').strftime('%d/%m/%Y')}"
        )
        y -= espacamento
        y = verificarEspaco(c, y, altura)

    y -= espacamento

    c.drawString(50, y, "Contas")
    y -= espacamento
    for conta in contas:
        c.drawString(
            50, y,
            f"{conta['descricao']} - R${conta['valor']} - Vencimento: {datetime.strptime(conta['vencimento'], '%Y-%m-%d').strftime('%d/%m/%Y')} - {conta['status']}"
        )
        y -= espacamento
        y = verificarEspaco(c, y, altura)

    c.save()
    return FileResponse(caminhoArquivo,
                        media_type="application/pdf",
                        filename="Relatorio.pdf")
